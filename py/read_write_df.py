import json
import os
import toml
from typing import Tuple, Union
import geopandas as gpd
import pandas as pd
import logging
import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo



DEBUG_MODE = True
logging.basicConfig(level=logging.DEBUG if DEBUG_MODE else logging.INFO)


class StatusTableManager:
    """Class to manage the metadata and formatting of a status table."""

    def __init__(self, metadata_file):
        self.metadata_file = metadata_file
        self.metadata = None

    def __enter__(self):
        """Load the metadata file when entering the context."""
        with open(self.metadata_file, 'r') as f:
            self.metadata = json.load(f)
        return self  # Return the instance to use in the context block

    def __exit__(self, exc_type, exc_value, traceback):
        """Clean up when exiting the context (if necessary)."""
        self._write_toml()  # Write the metadata back to the TOML file
        self.metadata = None  # Clear metadata from memory
        # Optional: Handle exceptions here if needed
        if exc_type:
            print(f"An exception of type {exc_type} occurred: {exc_value}")
        return False  # Returning False will propagate the exception, True will suppress it

    def _write_toml(self):
        json_to_toml(self.metadata_file)

    def get_column_names(self, target_format):
        """Get column names for a specific format."""
        return {key: value.get(target_format) for key, value in
                self.metadata["columns"].items()}

    def rename_columns(self, df, target_format, current_format):
        """
        Rename columns of a DataFrame from current_format to target_format.

        Args:
            df (pd.DataFrame): The input DataFrame.
            target_format (str): The desired output format (e.g., 'geojson', 'excel', 'shapefile').
            current_format (str): The current format of the input DataFrame.

        Returns:
            pd.DataFrame: The DataFrame with renamed columns.
        """
        # Reverse lookup to get universal keys from current column names
        reverse_map = {v[current_format]: k for k, v in self.metadata["columns"].items() if current_format in v}

        # Map current column names to universal keys
        universal_columns = {reverse_map.get(col, col): col for col in df.columns}

        # Map universal keys to target column names
        column_map = {key: self.metadata["columns"][key][target_format] for key in universal_columns if
                      key in self.metadata["columns"]}

        # Rename the DataFrame columns
        df = df.rename(columns={universal_columns[key]: column_map[key] for key in column_map})

        # Ensure all target columns are present, even if missing in input
        for key, target_col in column_map.items():
            if target_col not in df.columns:
                df[target_col] = None  # Default value for missing columns

        # Reorder columns to match the target format
        target_columns = [self.metadata["columns"][key][target_format] for key in self.metadata["columns"] if
                          target_format in self.metadata["columns"][key]]
        return df[[col for col in target_columns if col in df.columns]]

    def sort_rows(self, df):
        """Sort rows of a DataFrame by a specific column."""
        for col in self.metadata.get("sort_order", []):
            if col in df.columns:
                df = df.sort_values(by=col)
        return df

    def enforce_types(self, df, current_format="geojson"):
        """Enforce column data types based on metadata and current column names."""
        # Map current column names to metadata names
        column_map = {v[current_format]: k for k, v in self.metadata["columns"].items()}
        for current_col, metadata_col in column_map.items():
            if current_col in df.columns:  # Check if the column exists in the DataFrame
                tgt_dtype = self.metadata["columns"][metadata_col].get("dtype")

                if tgt_dtype == "date":
                    # Convert to date format without times
                    try:
                        df[current_col] = pd.to_datetime(df[current_col], errors="coerce")
                        df[current_col] = df[current_col].dt.strftime('%Y-%m-%d')
                        df[current_col] = df[current_col].fillna('')
                    except Exception as e:
                        logging.debug("Target dtype: %s", tgt_dtype)
                        print(f"Error processing column {current_col}: {e}")
                elif tgt_dtype == "string":
                    try:
                        # Convert to string
                        df[current_col] = df[current_col].astype(str)
                    except Exception as e:
                        logging.debug("Target dtype: %s", tgt_dtype)
                        print(f"Error processing column {current_col}: {e}")
                # Add other type conversions as needed (e.g., numeric)
                elif tgt_dtype == "numeric":
                    df[current_col] = pd.to_numeric(df[current_col], errors="coerce").fillna(0)
        return df


def df_to_excel(df: pd.DataFrame, out_loc: str, filename=None, sheetname="Sheet1"):

    value_sheetname = sheetname + "_values"

    print(f"\nExcel Stuff for {filename}")
    if filename is None:
        filename = "output"
    outpath = os.path.join(out_loc, f"{filename}.xlsx") if ".xlsx" not in filename else os.path.join(out_loc, filename)
    os.makedirs(out_loc, exist_ok=True)

    if isinstance(df, gpd.GeoDataFrame) or "geometry" in df.columns:
        df = pd.DataFrame(df.drop(columns=['geometry']))

    for c in df.columns:
        if "legend" in c.lower():
            df.drop(columns=c, inplace=True)

    # Check existing sheets for target and value sheets
    if os.path.exists(outpath):
        try:
            with pd.ExcelFile(outpath, engine="openpyxl") as reader:
                sheets = reader.sheet_names
                print(f" Sheet Names: {sheets}")
        except KeyError:
            os.remove(outpath)

    # Open the Excel file in append mode or create a new one
    with pd.ExcelWriter(outpath, engine="openpyxl", mode="w") as writer:
        if value_sheetname:
            # Ensure the target sheet exists and delete it safely
            if value_sheetname in writer.book.sheetnames:
                if len(writer.book.sheetnames) > 1:
                    writer.book.remove(writer.book[value_sheetname])
                else:
                    raise ValueError("Cannot delete the only visible sheet in the workbook.")
            df.to_excel(writer, index=False, sheet_name=value_sheetname)

        # Ensure at least one sheet is active and visible
        if writer.book.sheetnames:
            writer.book.active = 0  # Set the first sheet as active
            writer.book[writer.book.sheetnames[0]].sheet_state = "visible"

        if sheetname in writer.book.sheetnames:
            writer.book.remove(writer.book[sheetname])
        df.to_excel(writer, index=False, sheet_name=sheetname)

def df_to_excel_for_export(df: pd.DataFrame, out_loc: str, filename=None, sheetname="Sheet1", merged_headers=None):
    if filename is None:
        filename = "output.xlsx"
    else:
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

    outpath = os.path.join(out_loc, filename)
    os.makedirs(out_loc, exist_ok=True)

    if isinstance(df, gpd.GeoDataFrame) or "geometry" in df.columns:
        df = df.drop(columns='geometry')

    # Remove legend columns if present
    for c in df.columns:
        if "legend" in c.lower():
            df.drop(columns=c, inplace=True)

    # Create a new workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheetname

    # Apply merged headers if provided
    if merged_headers:
        # Create a Font object with the desired font name
        custom_font = Font(name='Century Gothic', size=12, bold=True, color='F7F7F7')
        dark_background_fill = PatternFill(start_color='000033', end_color='000033', fill_type='solid')
        current_col = 1
        for header in merged_headers:
            start_col = current_col
            colspan = header.get('colspan', 1)
            end_col = current_col + colspan - 1
            cell = ws.cell(row=1, column=start_col)
            cell.value = header['text']
            cell.font = custom_font
            cell.fill = dark_background_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if colspan > 1:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

            current_col += colspan

        # Write column headers starting from row 2
        for idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=2, column=idx)
            cell.value = col_name
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        start_row = 3  # Data starts from row 3
    else:
        # No merged headers; write column headers in row 1
        for idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.value = col_name
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        start_row = 2  # Data starts from row 2

    # Write data to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Freeze panes (freeze top rows)
    freeze_row = start_row if merged_headers else start_row - 1
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)

    # Adjust column widths (optional)
    for column_cells in ws.columns:
        # Get the column letter using the column index
        col_idx = column_cells[0].column  # Column index (integer)
        col_letter = get_column_letter(col_idx)
        # Calculate the maximum length needed
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[col_letter].width = length + 2

    ws = format_as_excel_table(ws, start_row - 1 if merged_headers else start_row - 2)

    # Save the workbook
    wb.save(outpath)
    print(f"Excel file saved to {outpath}")


def format_as_excel_table(worksheet, table_start_row=1):
    # Create an Excel Table
    start_row = table_start_row
    table_end_row = worksheet.max_row
    table_start_col = 1
    table_end_col = worksheet.max_column
    table_ref = f"{worksheet.cell(row=start_row, column=table_start_col).coordinate}:{worksheet.cell(row=table_end_row, column=table_end_col).coordinate}"

    tab = Table(displayName="DataTable", ref=table_ref)

    # Add a table style
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    worksheet.add_table(tab)

    return worksheet


def df_to_json(data, out_loc: str, filename: str = None):
    if isinstance(data, gpd.GeoDataFrame):
        df = pd.DataFrame(data.drop(columns='geometry'))
    elif isinstance(data, pd.DataFrame):
        df = data
    else:
        raise ValueError("Data must be a GeoDataFrame or DataFrame")

    if ".json" in out_loc:
        out_loc, file = os.path.split(out_loc)
        filename, ext = os.path.splitext(file)
    outpath_table = os.path.normpath(os.path.join(out_loc, filename + ".json"))

    # Convert Timestamp and date objects to strings
    dicted = df.to_dict(orient='records')

    with open(outpath_table, 'w') as f:
        json.dump(dicted, f, indent=2)


def df_to_metadata(data, out_loc: str, filename: str = None):
    # Ensure data is a GeoDataFrame or DataFrame
    if isinstance(data, gpd.GeoDataFrame):
        df = pd.DataFrame(data.drop(columns='geometry'))  # Drop geometry for GeoDataFrame
    elif isinstance(data, pd.DataFrame):
        df = data.copy()
    else:
        raise ValueError("Data must be a GeoDataFrame or DataFrame")

    # Automatically convert datetime columns to strings
    for col in df.select_dtypes(include=['datetime', 'datetime64[ns]']).columns:
        df[col] = df[col].dt.strftime('%Y-%m-%d')  # Format as 'YYYY-MM-DD'

    # Ensure all other types are safely converted into string if required
    for col in df.columns:
        if df[col].dtype not in [bool, int, float, 'datetime64[ns]'] and not pd.api.types.is_numeric_dtype(df[col]):
            df[col] = df[col].astype(str)

    # Handle output file path
    if ".json" in out_loc:  # If file path is given directly
        out_loc, file = os.path.split(out_loc)
        filename, ext = os.path.splitext(file)
    if filename is None:  # Default filename if not provided
        filename = "metadata"
    outpath_table = os.path.normpath(os.path.join(out_loc, filename + ".json"))

    # Convert DataFrame to JSON records
    dicted = df.to_dict(orient='records')

    # Convert to unique values
    unique_dict = {}
    for row in dicted:
        for key, value in row.items():
            if key not in unique_dict:
                unique_dict[key] = [value]
            if value not in unique_dict[key]:
                unique_dict[key].append(value)
    unique_dict = {k: list(set(v)) for k, v in unique_dict.items()}

    # Write to JSON file
    os.makedirs(out_loc, exist_ok=True)  # Ensure directory exists
    with open(outpath_table, 'w') as f:
        json.dump(unique_dict, f, indent=2)

    print(f"Metadata successfully written to {outpath_table}")


def gdf_to_geojson(gdf: gpd.GeoDataFrame, out_loc, filename=None):
    if not isinstance(gdf, gpd.GeoDataFrame):
        if isinstance(gdf, gpd.GeoSeries):
            gdf = gpd.GeoDataFrame(geometry=gdf)
        print(f"Data type: {type(gdf)}")
        raise ValueError("Data must be a GeoDataFrame")
    driver = "GeoJSON"
    outpath = out_loc + f"{filename}.geojson"
    os.makedirs(out_loc, exist_ok=True)
    gdf.to_file(filename=outpath, driver=driver)

    print(f"Saved {filename} to {outpath}")
    print(f"Columns: {gdf.columns.to_list()}")


def gdf_to_shapefile(gdf: gpd.GeoDataFrame, out_loc):
    """
    Save a GeoDataFrame to a shapefile
    :param gdf:
    :param out_loc:
    :return:
    """

    if not isinstance(gdf, gpd.GeoDataFrame):
        if isinstance(gdf, gpd.GeoSeries):
            gdf = gpd.GeoDataFrame(geometry=gdf)
        print(f"Data type: {type(gdf)}")
        raise ValueError("Data must be a GeoDataFrame")

    driver = "ESRI Shapefile"

    if ".shp" in out_loc:
        outpath = out_loc
        out_loc, filename = os.path.split(out_loc)
    else:
        filename = "OUTPUT_FILE.shp"
        outpath = out_loc + f"OUTPUT_FILE.shp"
    os.makedirs(out_loc, exist_ok=True)
    gdf.to_file(outpath, driver=driver)

    print(f"Saved {filename} to {outpath}")


def json_to_toml(input_json: Union[json, str], toml_file: str = None
                 ) -> Tuple[dict, str]:
    """
    Convert a JSON file to a TOML file
    :param input_json:
    :param toml_file:
    :return: json data, toml file path
    """
    if isinstance(input_json, str):
        with open(input_json, "r") as jf:
            data = json.load(jf)
    else:
        if not toml_file:
            raise ValueError("Must give output TOML file path")
        data = input_json

    # Export to TOML
    if not toml_file:
        toml_file = input_json.replace(".json", ".toml")
    with open(toml_file, "w") as tf:
        toml.dump(data, tf)

    return data, toml_file


def toml_to_json(toml_file: str, json_file: str = None
                 ) -> Tuple[dict, str]:
    """
    Convert a TOML file to a JSON file
    :param toml_file:
    :param json_file:
    :return: json data, json file path
    """
    with open(toml_file, "r") as tf:
        data = toml.load(tf)

    # Export to JSON
    if not json_file:
        json_file = toml_file.replace(".toml", ".json")
    with open(json_file, "w") as jf:
        json.dump(data, jf, indent=2)

    return data, json_file


if __name__ == "__main__":
    pass
