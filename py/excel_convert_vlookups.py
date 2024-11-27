import openpyxl


def convert_vlookups_to_values(workbook_path="Iowa_BLE_Tracking.xlsx",
                               vlookup_sheet_name="Tracking_VLOOKUP",
                               new_sheet_name="Tracking_Evaluated"):
    # Load the workbook
    wb = openpyxl.load_workbook(workbook_path, data_only=True)  # `data_only=True` gets calculated values

    # Access the sheet with formulas
    sheet = wb[vlookup_sheet_name]

    # Extract the evaluated values into a list of lists
    calculated_values = []
    for row in sheet.iter_rows(values_only=True):  # `values_only=True` returns cell values, not formulas
        calculated_values.append(list(row))

    # Optional: Create a new sheet with static values
    if new_sheet_name in wb.sheetnames:
        del wb[new_sheet_name]  # Remove existing sheet if it exists
    wb.create_sheet(new_sheet_name)
    new_sheet = wb[new_sheet_name]

    # Write the calculated values to the new sheet
    for row_idx, row_values in enumerate(calculated_values, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            new_sheet.cell(row=row_idx, column=col_idx, value=value)

    # Save the workbook with the new sheet
    wb.save(workbook_path)
